"""Format-neutral Office normalization with original-file provenance."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


OOXML = {".docx", ".pptx", ".xlsx"}
IMAGES = {".png": "png", ".jpg": "jpeg", ".jpeg": "jpeg", ".webp": "webp"}
LEGACY_TARGET = {".doc": "docx", ".ppt": "pptx", ".xls": "xlsx"}


def source_path(path: Path) -> Path:
    marker = ".proofline."
    return path.with_name(path.name.split(marker, 1)[0]) if marker in path.name else path


def normalize_legacy(root: Path) -> int:
    """Convert binary Office files when LibreOffice is available; originals remain authoritative."""
    executable = shutil.which("soffice")
    if not executable:
        return 0
    converted = 0
    for path in list(root.rglob("*")):
        target = LEGACY_TARGET.get(path.suffix.casefold()) if path.is_file() else None
        derived = path.with_name(f"{path.name}.proofline.{target}") if target else None
        if not target or derived.exists():
            continue
        with tempfile.TemporaryDirectory(dir=root) as output:
            result = subprocess.run(
                [executable, "--headless", "--convert-to", target, "--outdir", output, str(path)],
                capture_output=True,
                timeout=60,
                check=False,
            )
            generated = Path(output) / path.with_suffix(f".{target}").name
            if result.returncode == 0 and generated.exists():
                generated.replace(derived)
                converted += 1
    return converted


def _xml(path: Path, names: list[str]) -> list[tuple[str, ET.Element]]:
    try:
        with zipfile.ZipFile(path) as package:
            return [(name, ET.fromstring(package.read(name))) for name in names if name in package.namelist()]
    except (zipfile.BadZipFile, KeyError, ET.ParseError):
        return []


def pptx_passages(path: Path) -> list[tuple[str, str, dict]]:
    try:
        with zipfile.ZipFile(path) as package:
            names = sorted(
                (name for name in package.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")),
                key=lambda name: int(Path(name).stem.removeprefix("slide")),
            )
    except zipfile.BadZipFile:
        return []
    passages = []
    for slide, (_, tree) in enumerate(_xml(path, names), 1):
        text = " ".join(node.text or "" for node in tree.iter() if node.tag.endswith("}t")).strip()
        if text:
            passages.append((text, "slide", {"slide": slide}))
    return passages


def docx_passages(path: Path) -> list[tuple[str, str, dict]]:
    trees = _xml(path, ["word/document.xml"])
    if not trees:
        return []
    paragraphs = []
    for node in trees[0][1].iter():
        if not node.tag.endswith("}p"):
            continue
        text = "".join(child.text or "" for child in node.iter() if child.tag.endswith("}t")).strip()
        if text:
            paragraphs.append((text, "paragraph", {"paragraph": len(paragraphs) + 1}))
    return paragraphs


def office_tables(root: Path) -> list[dict]:
    tables: list[dict] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        suffix = path.suffix.casefold()
        original = source_path(path)
        relative = original.relative_to(root).as_posix()
        if suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                    if rows:
                        tables.append({"file": relative, "sheet": sheet.title, "rows": rows, "path": original})
            finally:
                workbook.close()
        elif suffix in {".docx", ".pptx"}:
            prefix = "word/" if suffix == ".docx" else "ppt/slides/"
            try:
                with zipfile.ZipFile(path) as package:
                    names = [name for name in package.namelist() if name.startswith(prefix) and name.endswith(".xml")]
            except zipfile.BadZipFile:
                continue
            number = 0
            for name, tree in _xml(path, names):
                for table in (node for node in tree.iter() if node.tag.endswith("}tbl")):
                    rows = []
                    for row in (node for node in table if node.tag.endswith("}tr")):
                        cells = [
                            " ".join(text.text or "" for text in cell.iter() if text.tag.endswith("}t")).strip()
                            for cell in row if cell.tag.endswith("}tc")
                        ]
                        if cells:
                            rows.append(cells)
                    if rows:
                        number += 1
                        location = f"Slide {Path(name).stem.removeprefix('slide')}" if suffix == ".pptx" else "Document"
                        tables.append({"file": relative, "sheet": f"{location} Table {number}", "rows": rows, "path": original})
        elif suffix == ".md":
            blocks = re.split(r"\n\s*\n", path.read_text(encoding="utf-8", errors="replace"))
            for number, block in enumerate(blocks, 1):
                rows = [[cell.strip() for cell in line.strip().strip("|").split("|")] for line in block.splitlines() if "|" in line]
                rows = [row for row in rows if not all(re.fullmatch(r":?-{3,}:?", cell or "") for cell in row)]
                if len(rows) > 1:
                    tables.append({"file": relative, "sheet": f"Markdown Table {number}", "rows": rows, "path": original})
        elif suffix == ".json":
            try:
                value = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, UnicodeError, json.JSONDecodeError):
                continue
            groups = [("JSON", value)] if isinstance(value, list) else [(str(key), rows) for key, rows in value.items() if isinstance(rows, list)] if isinstance(value, dict) else []
            for name, records in groups:
                if records and all(isinstance(record, dict) for record in records):
                    headers = list(dict.fromkeys(key for record in records for key in record))
                    tables.append({"file": relative, "sheet": name, "rows": [headers, *[[record.get(key) for key in headers] for record in records]], "path": original})
        elif suffix == ".xml":
            try:
                tree = ET.parse(path).getroot()
            except (OSError, ET.ParseError):
                continue
            groups: dict[str, list[dict]] = {}
            for node in tree.iter():
                record = {Path(child.tag).name.split("}")[-1]: (child.text or "").strip() for child in node if len(child) == 0}
                if len(record) > 1:
                    groups.setdefault(node.tag.split("}")[-1], []).append(record)
            for name, records in groups.items():
                headers = list(dict.fromkeys(key for record in records for key in record))
                tables.append({"file": relative, "sheet": name, "rows": [headers, *[[record.get(key) for key in headers] for record in records]], "path": original})
    return tables


def office_media(root: Path) -> list[dict]:
    media: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.casefold() not in OOXML:
            continue
        original = source_path(path)
        try:
            with zipfile.ZipFile(path) as package:
                for name in package.namelist():
                    suffix = Path(name).suffix.casefold()
                    if "/media/" not in name or suffix not in IMAGES:
                        continue
                    data = package.read(name)
                    identity = (str(original), hashlib.sha256(data).hexdigest())
                    if identity in seen:
                        continue
                    seen.add(identity)
                    media.append({
                        "path": original,
                        "data": data,
                        "media": IMAGES[suffix],
                        "locator": {"embedded_object": Path(name).name},
                    })
        except (zipfile.BadZipFile, KeyError):
            continue
    return media
