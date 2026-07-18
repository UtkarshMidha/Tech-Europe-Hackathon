"""AI fallback for tabular sources the deterministic schema discovery cannot map."""

from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Annotated, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from backend.ingest import office_tables


ShortText = Annotated[str, Field(max_length=160)]
Role = Literal[
    "general_ledger", "vendor_transactions", "vendor_master", "accounts",
    "assets", "asset_postings", "goods_receipts", "master_changes",
    "supplier_invoice_journal", "sales_invoice_journal", "permissions", "other",
]
Target = Literal[
    "SACHKONTONUMMER", "SACHKONTONAME", "KONTENART", "BUCHUNGSBETRAG",
    "BUCHUNGSDATUM", "BUCHUNGSNUMMER", "BUCHUNGSTEXT", "BUCHUNGSTYP",
    "DOKUMENT", "BELEGNUMMER", "BENUTZERKENNUNG", "LIEFERANTENKONTONUMMER",
    "LIEFERANTENNAME", "ANLAGENNUMMER", "ANLAGENBEZEICHNUNG", "ANLAGENGRUPPE",
    "BUCHUNGSART", "KREDITOR", "KREDITORNAME", "WARENEINGANG_DATUM",
    "RECHNUNGSNUMMER", "BETRAG_EUR", "BEMERKUNG", "KONTO", "FELD",
    "GEAENDERT_VON", "GENEHMIGT_VON", "GENEHMIGT", "DATUM", "WERT_ALT",
    "WERT_NEU", "FAKTURADATUM", "LEISTUNGSDATUM", "DEBITOR", "DEBITORNAME",
    "Benutzer", "Buchen", "Zahlungslauf", "Stammdaten/Kreditor anlegen",
]


class ColumnMapping(BaseModel):
    source: ShortText
    target: Target


class SourceMapping(BaseModel):
    file: ShortText
    sheet: ShortText | None = None
    role: Role
    header_row: int = Field(ge=1, le=20)
    confidence: Literal["low", "medium", "high"]
    columns: list[ColumnMapping] = Field(max_length=30)


class MappingPlan(BaseModel):
    sources: list[SourceMapping] = Field(max_length=30)


REQUIRED = {
    "general_ledger": {"SACHKONTONUMMER", "BUCHUNGSBETRAG", "DOKUMENT", "BUCHUNGSDATUM", "BENUTZERKENNUNG"},
    "vendor_transactions": {"LIEFERANTENKONTONUMMER", "BUCHUNGSNUMMER", "BUCHUNGSBETRAG", "BUCHUNGSTEXT"},
    "vendor_master": {"LIEFERANTENKONTONUMMER", "LIEFERANTENNAME"},
    "accounts": {"SACHKONTONUMMER", "SACHKONTONAME"},
    "assets": {"ANLAGENNUMMER", "ANLAGENBEZEICHNUNG", "ANLAGENGRUPPE"},
    "asset_postings": {"ANLAGENNUMMER", "BELEGNUMMER", "BUCHUNGSART"},
    "goods_receipts": {"KREDITOR", "WARENEINGANG_DATUM", "BETRAG_EUR"},
    "master_changes": {"KONTO", "FELD", "GEAENDERT_VON", "GENEHMIGT_VON"},
    "supplier_invoice_journal": {"RECHNUNGSNUMMER", "KREDITOR", "FAKTURADATUM", "LEISTUNGSDATUM", "BETRAG_EUR"},
    "sales_invoice_journal": {"RECHNUNGSNUMMER", "DEBITOR", "FAKTURADATUM", "LEISTUNGSDATUM", "BETRAG_EUR"},
    "permissions": {"Benutzer", "Buchen", "Zahlungslauf", "Stammdaten/Kreditor anlegen"},
}


def _text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def _delimited_rows(path: Path) -> tuple[str, list[list[str]]]:
    text = _text(path)
    sample = text[:8192]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        delimiter = ";"
    return delimiter, list(csv.reader(text.splitlines(), delimiter=delimiter))


def inventory(root: Path) -> list[dict]:
    items: list[dict] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        relative = path.relative_to(root).as_posix()
        if path.suffix.casefold() in {".csv", ".txt"} and path.stat().st_size <= 10 * 1024 * 1024:
            _, rows = _delimited_rows(path)
            if rows:
                items.append({"file": relative, "sheet": None, "preview": [row[:20] for row in rows[:7]]})
    for table in office_tables(root):
        items.append({"file": table["file"], "sheet": table["sheet"], "preview": [row[:20] for row in table["rows"][:10]]})
    return items


def _source_rows(root: Path, mapping: SourceMapping) -> tuple[Path, list[list[object]]]:
    path = (root / mapping.file).resolve()
    if root.resolve() not in path.parents or not path.is_file():
        return path, []
    if path.suffix.casefold() in {".csv", ".txt"}:
        return path, _delimited_rows(path)[1]
    if path.suffix.casefold() == ".xlsx" and mapping.sheet:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if mapping.sheet not in workbook.sheetnames:
                return path, []
            return path, [list(row) for row in workbook[mapping.sheet].iter_rows(values_only=True)]
        finally:
            workbook.close()
    table = next((item for item in office_tables(root) if item["file"] == mapping.file and item["sheet"] == mapping.sheet), None)
    if table:
        return table["path"], table["rows"]
    return path, []


def _materialize(root: Path, plan: MappingPlan) -> tuple[dict[str, list[dict]], list[dict]]:
    tables: dict[str, list[dict]] = {}
    accepted: list[dict] = []
    for mapping in plan.sources:
        targets = {column.target for column in mapping.columns}
        if mapping.role == "other" or mapping.confidence == "low" or not REQUIRED.get(mapping.role, set()) <= targets:
            continue
        path, rows = _source_rows(root, mapping)
        if len(rows) <= mapping.header_row:
            continue
        headers = [str(value or "").strip() for value in rows[mapping.header_row - 1]]
        indexes = {header: index for index, header in enumerate(headers)}
        if not all(column.source in indexes for column in mapping.columns):
            continue
        records = []
        for row_number, row in enumerate(rows[mapping.header_row:], start=mapping.header_row + 1):
            record = {
                column.target: row[indexes[column.source]] if indexes[column.source] < len(row) else None
                for column in mapping.columns
            }
            if any(value not in (None, "") for value in record.values()):
                records.append({**record, "_row": row_number, "_file": path, "_sheet": mapping.sheet, "_role": mapping.role})
        if records:
            key = f"adaptive:{mapping.file}:{mapping.sheet or ''}"
            tables[key] = records
            accepted.append(mapping.model_dump())
    return tables, accepted


def map_tables(root: Path) -> tuple[dict[str, list[dict]], list[dict]]:
    """Ask AI only after native schema discovery has failed."""
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return {}, []
    from openai import OpenAI

    client = OpenAI(api_key=api_key, timeout=60.0, max_retries=0)
    tables: dict[str, list[dict]] = {}
    accepted: list[dict] = []
    sources = inventory(root)
    for start in range(0, len(sources), 25):
        response = client.responses.parse(
            model=os.getenv("OPENAI_MODEL", "gpt-5.6"),
            instructions=(
                "Map unfamiliar audit tables into the supplied canonical accounting fields. Use headers and sample values, "
                "preserve source meaning, and return other/low confidence when uncertain. Do not invent columns. "
                "A general ledger must map account, amount, document, posting date, and user fields."
            ),
            input=json.dumps({"sources": sources[start:start + 25]}, ensure_ascii=False, default=str),
            text_format=MappingPlan,
            reasoning={"effort": "low"},
            max_output_tokens=5000,
            store=False,
            safety_identifier="proofline-demo",
            text={"verbosity": "low"},
        )
        if response.output_parsed:
            mapped, mappings = _materialize(root, response.output_parsed)
            tables.update(mapped)
            accepted.extend(mappings)
    return tables, accepted
