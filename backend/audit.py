"""Deterministic dossier ingestion and evidence-backed audit detectors."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from backend.ingest import docx_passages, pptx_passages, source_path


BLOCKED_NAMES = ("GROUND-TRUTH", "SEALED", "ANSWER")
REPAIR_WORDS = (
    "reparatur", "instandsetzung", "austausch", "generalüberholung", "überholung",
    "repair", "maintenance", "replacement", "overhaul",
)

HEADER_ALIASES = {
    "VENDORACCOUNT": "LIEFERANTENKONTONUMMER", "VENDOR_ACCOUNT": "LIEFERANTENKONTONUMMER",
    "VENDOR_ID": "LIEFERANTENKONTONUMMER", "SUPPLIER_ID": "LIEFERANTENKONTONUMMER",
    "VENDORNAME": "LIEFERANTENNAME", "VENDOR_NAME": "LIEFERANTENNAME", "SUPPLIER_NAME": "LIEFERANTENNAME",
    "MAINACCOUNT": "SACHKONTONUMMER", "MAIN_ACCOUNT": "SACHKONTONUMMER", "GL_ACCOUNT": "SACHKONTONUMMER",
    "ACCOUNT_NAME": "SACHKONTONAME", "GL_ACCOUNT_NAME": "SACHKONTONAME",
    "TRANSACTIONAMOUNT": "BUCHUNGSBETRAG", "TRANSACTION_AMOUNT": "BUCHUNGSBETRAG", "BOOKING_AMOUNT": "BUCHUNGSBETRAG",
    "AMOUNT_EUR": "BETRAG_EUR", "POSTING_DATE": "BUCHUNGSDATUM", "TRANSACTION_DATE": "BUCHUNGSDATUM",
    "VOUCHER": "BUCHUNGSNUMMER", "VOUCHER_ID": "BUCHUNGSNUMMER", "DOCUMENT_ID": "DOKUMENT", "DOCUMENT": "DOKUMENT", "DOCUMENT_NO": "DOKUMENT",
    "DOCUMENT_NUMBER": "BELEGNUMMER", "EXTERNAL_INVOICE": "BELEGNUMMER", "INVOICE_NUMBER": "RECHNUNGSNUMMER",
    "TRANSACTION_TEXT": "BUCHUNGSTEXT", "DESCRIPTION": "BUCHUNGSTEXT", "TRANSACTION_TYPE": "BUCHUNGSTYP", "USER_ID": "BENUTZERKENNUNG",
    "INVOICE_DATE": "FAKTURADATUM", "PERFORMANCE_DATE": "LEISTUNGSDATUM", "SERVICE_DATE": "LEISTUNGSDATUM",
    "GOODS_RECEIPT_DATE": "WARENEINGANG_DATUM", "RECEIPT_DATE": "WARENEINGANG_DATUM",
    "CHANGED_BY": "GEAENDERT_VON", "APPROVED_BY": "GENEHMIGT_VON", "FIELD": "FELD",
    "ASSET_ID": "ANLAGENNUMMER", "ASSET_NAME": "ANLAGENBEZEICHNUNG", "ASSET_GROUP": "ANLAGENGRUPPE",
    "USER": "Benutzer", "POST": "Buchen", "PAYMENT_RUN": "Zahlungslauf", "CREATE_VENDOR": "Stammdaten/Kreditor anlegen",
    "ACCOUNT": "Konto", "ACCOUNT_DESCRIPTION": "Bezeichnung", "ACCOUNT_TYPE": "Kontenart",
}


def _header(value: object) -> str:
    raw = str(value or "").strip()
    normalized = re.sub(r"[^A-Z0-9ÄÖÜ_]+", "_", raw.upper()).strip("_")
    return HEADER_ALIASES.get(normalized, raw)


def money(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    raw = str(value).strip().replace("€", "").replace("EUR", "").replace(" ", "")
    negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()")
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", "") if re.fullmatch(r"-?\d{1,3}(?:,\d{3})+", raw) else raw.replace(",", ".")
    elif raw.count(".") > 1 and re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", raw):
        raw = raw.replace(".", "")
    try:
        result = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid monetary value: {value!r}") from exc
    if not result.is_finite():
        raise ValueError(f"Invalid monetary value: {value!r}")
    return -abs(result) if negative else result


def _find_rows(collection: dict[str, list[dict]], required: set[str]) -> list[dict]:
    """Select a table by schema, so renamed final-dossier files still work."""
    candidates = [rows for rows in collection.values() if rows and required <= set(rows[0])]
    return max(candidates, key=len, default=[])


def _date(value: object) -> date | None:
    raw = str(value or "").strip()
    for pattern in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw[:10], pattern).date()
        except ValueError:
            pass
    return None


def _enabled(value: object) -> bool:
    return str(value or "").strip().casefold() in {"x", "yes", "ja", "true", "1", "enabled"}


def _xlsx_table(
    workbooks: dict[str, dict[str, list[dict]]], required: set[str]
) -> tuple[Path | None, str | None, list[dict]]:
    """Find and normalize the first worksheet carrying the requested header set."""
    best: tuple[int, Path, str, list[dict]] | None = None
    for sheets in workbooks.values():
        for sheet_name, rows in sheets.items():
            for header_row in rows[:12]:
                headers = [_header(value) for value in header_row["values"]]
                if not required <= set(headers):
                    continue
                normalized = []
                for row in rows[header_row["_row"]:]:
                    values = list(row["values"])
                    if not any(value not in (None, "") for value in values):
                        continue
                    normalized.append({
                        **dict(zip(headers, values)),
                        "_row": row["_row"],
                        "_file": row["_file"],
                        "_sheet": sheet_name,
                    })
                candidate = (len(normalized), Path(header_row["_file"]), sheet_name, normalized)
                if best is None or candidate[0] > best[0]:
                    best = candidate
    return (best[1], best[2], best[3]) if best else (None, None, [])


def _encoding(path: Path) -> str:
    data = path.read_bytes()
    for candidate in ("utf-8-sig", "cp1252"):
        try:
            text = data.decode(candidate)
            if "�" not in text:
                return candidate
        except UnicodeDecodeError:
            pass
    return "cp1252"


def _csv_rows(path: Path, headers: list[str] | None = None) -> list[dict]:
    text = path.read_text(encoding=_encoding(path))
    try:
        delimiter = ";" if headers is not None else csv.Sniffer().sniff(text[:8192], delimiters=";,\t|").delimiter
    except csv.Error:
        delimiter = ";"
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter, quotechar='"'))
    if not rows:
        return []
    offset = 1
    if headers is None:
        headers, rows, offset = rows[0], rows[1:], 2
    headers = [_header(header) for header in headers]
    return [
        {**dict(zip(headers, row)), "_row": index, "_file": path}
        for index, row in enumerate(rows, start=offset)
    ]


def _gdpdu_tables(root: Path) -> dict[str, list[dict]]:
    tables: dict[str, list[dict]] = {}
    for index in root.rglob("index.xml"):
        try:
            xml = ET.parse(index).getroot()
        except ET.ParseError:
            continue
        for table in xml.findall(".//Table"):
            url = table.findtext("URL")
            columns = [node.findtext("Name", "") for node in table.findall(".//VariableColumn")]
            if not url or not columns:
                continue
            path = index.parent / url
            if path.exists():
                tables[url] = _csv_rows(path, columns)
    return tables


def _companion_csv(root: Path) -> dict[str, list[dict]]:
    return {path.name: _csv_rows(path) for path in root.rglob("*.csv")}


def _docx_text(path: Path) -> list[str]:
    try:
        with zipfile.ZipFile(path) as package:
            xml = ET.fromstring(package.read("word/document.xml"))
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        return [
            "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
            for paragraph in xml.findall(".//w:p", ns)
            if "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
        ]
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        return []


def _pdf_pages(path: Path) -> list[str]:
    try:
        from pypdf import PdfReader

        return [(page.extract_text() or "").strip() for page in PdfReader(path).pages]
    except Exception:
        return []


def _document_passages(path: Path) -> list[tuple[str, str, dict]]:
    """Return searchable policy passages with source-native locators."""
    if path.suffix.casefold() == ".docx":
        return docx_passages(path)
    if path.suffix.casefold() == ".pptx":
        return pptx_passages(path)
    if path.suffix.casefold() == ".pdf":
        return [
            (line.strip(), "page", {"page": page})
            for page, text in enumerate(_pdf_pages(path), 1)
            for line in text.splitlines()
            if line.strip()
        ]
    if path.suffix.casefold() in {".txt", ".csv", ".xml", ".md", ".json"}:
        try:
            return [(text.strip(), "row", {"row": row}) for row, text in enumerate(path.read_text(encoding=_encoding(path)).splitlines(), 1) if text.strip()]
        except (OSError, UnicodeError):
            return []
    return []


def _xlsx_rows(path: Path) -> dict[str, list[dict]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        result: dict[str, list[dict]] = {}
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            result[sheet.title] = [
                {"_row": index, "values": list(row), "_file": path}
                for index, row in enumerate(rows, start=1)
            ]
        return result
    finally:
        workbook.close()


def _trial_balance(path: Path | None) -> dict:
    empty = {
        "reported_profit": None,
        "formula_count": 0,
        "formula_resolved": 0,
        "sheet": None,
        "reported_sheet": None,
        "reported_cell": None,
    }
    if not path:
        return empty
    workbook = load_workbook(path, data_only=False)
    sheet = None
    header_row = 0
    headers: dict[str, int] = {}
    for candidate in workbook.worksheets:
        for row in range(1, min(candidate.max_row, 12) + 1):
            values = [_header(candidate.cell(row, column).value) for column in range(1, candidate.max_column + 1)]
            if {"Konto", "Bezeichnung", "Kontenart"} <= set(values):
                sheet, header_row = candidate, row
                headers = {value: index for index, value in enumerate(values, start=1) if value}
                break
        if sheet:
            break
    if not sheet:
        workbook.close()
        return empty

    opening_column = next((column for name, column in headers.items() if name.casefold().startswith(("eb ", "opening"))), None)
    debit_column = next((column for name, column in headers.items() if "soll" in name.casefold() or "debit" in name.casefold()), None)
    credit_column = next((column for name, column in headers.items() if "haben" in name.casefold() or "credit" in name.casefold()), None)
    balance_column = next((column for name, column in headers.items() if "saldo" in name.casefold() or "balance" in name.casefold()), None)
    type_column = headers["Kontenart"]
    if not all((opening_column, debit_column, credit_column, balance_column)):
        workbook.close()
        return empty

    balances: dict[int, Decimal] = {}
    formula_count = 0
    formula_resolved = 0
    data_rows: list[int] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        account_type = str(sheet.cell(row, type_column).value or "").strip()
        if account_type.casefold() not in {"guv", "p&l", "profit and loss", "income statement", "bilanz", "balance sheet"}:
            continue
        data_rows.append(row)
        opening = money(sheet.cell(row, opening_column).value)
        debit = money(sheet.cell(row, debit_column).value)
        credit = money(sheet.cell(row, credit_column).value)
        balances[row] = opening + debit - credit
        formula = sheet.cell(row, balance_column).value
        if isinstance(formula, str) and formula.startswith("="):
            formula_count += 1
            expected = (
                f"={sheet.cell(row, opening_column).coordinate}+"
                f"{sheet.cell(row, debit_column).coordinate}-"
                f"{sheet.cell(row, credit_column).coordinate}"
            )
            formula_resolved += int(formula.replace("$", "").replace(" ", "").upper() == expected.upper())

    # Validate total formulas explicitly; do not count arbitrary spreadsheet formulas as resolved.
    if data_rows:
        total_row = max(data_rows) + 1
        for column in (opening_column, debit_column, credit_column):
            formula = sheet.cell(total_row, column).value
            if isinstance(formula, str) and formula.startswith("="):
                formula_count += 1
                letter = sheet.cell(1, column).column_letter
                expected = f"=SUM({letter}{min(data_rows)}:{letter}{max(data_rows)})"
                formula_resolved += int(formula.replace("$", "").replace(" ", "").upper() == expected.upper())

    profit = -sum(
        balances[row]
        for row in data_rows
        if str(sheet.cell(row, type_column).value or "").strip().casefold() in {"guv", "p&l", "profit and loss", "income statement"}
    )
    reported_sheet = None
    reported_cell = None
    for bridge in workbook.worksheets:
        if bridge is sheet:
            continue
        for row in range(1, min(bridge.max_row, 20) + 1):
            for column in range(1, bridge.max_column + 1):
                label = str(bridge.cell(row, column).value or "").casefold()
                if "jahresüberschuss laut entwurf" in label or "draft profit" in label:
                    candidate = bridge.cell(row, column + 1)
                    if candidate.value not in (None, ""):
                        profit = money(candidate.value)
                        reported_sheet, reported_cell = bridge.title, candidate.coordinate
                formula = bridge.cell(row, column).value
                if isinstance(formula, str) and formula.startswith("=") and "SUMIF" in formula.upper():
                    formula_count += 1
                    # The supported bridge formula sums the computed P&L balance column.
                    formula_resolved += int(sheet.title.casefold() in formula.casefold() and any(term in formula.casefold() for term in ("guv", "p&l", "profit and loss")))
    result = {
        "reported_profit": profit,
        "formula_count": formula_count,
        "formula_resolved": formula_resolved,
        "sheet": sheet.title,
        "reported_sheet": reported_sheet or sheet.title,
        "reported_cell": reported_cell,
    }
    workbook.close()
    return result


class EvidenceBook:
    def __init__(self, root: Path):
        self.root = root
        self.items: dict[str, dict] = {}

    def add(self, path: Path, kind: str, locator: dict, label: str, excerpt: str, fields: dict | None = None) -> str:
        path = source_path(path)
        digest = hashlib.sha256(path.read_bytes()).hexdigest() if path.exists() else ""
        identity = json.dumps([str(path), locator, label], sort_keys=True, ensure_ascii=False).encode("utf-8")
        evidence_id = "ev-" + hashlib.sha1(identity).hexdigest()[:10]
        try:
            relative = path.relative_to(self.root).as_posix()
        except ValueError:
            relative = path.name
        self.items[evidence_id] = {
            "id": evidence_id,
            "file": relative,
            "kind": kind,
            "locator": locator,
            "label": label,
            "excerpt": excerpt,
            "fields": fields or {},
            "sha256": digest,
        }
        return evidence_id


def _record_excerpt(record: dict, keys: list[str]) -> str:
    return " · ".join(f"{key}: {record.get(key, '')}" for key in keys if record.get(key) not in (None, ""))


def _row_range(records: list[dict]) -> dict:
    rows = [int(record["_row"]) for record in records]
    return {"row_start": min(rows), "row_end": max(rows)} if rows else {"row_start": 0, "row_end": 0}


def _graph(nodes: list[tuple[str, str, str]], edges: list[tuple[str, str, str]]) -> dict:
    return {
        "nodes": [{"id": node, "label": label, "type": kind, "meta": {}} for node, label, kind in nodes],
        "edges": [{"source": source, "target": target, "label": label} for source, target, label in edges],
    }


def _hash_verification(root: Path, source_files: list[Path]) -> dict:
    """Verify exported source hashes against any supplied chain-of-custody PDF."""
    protocol_text = "".join(
        re.sub(r"\s+", "", page).casefold()
        for path in root.rglob("*.pdf")
        for page in _pdf_pages(path)
        if "sha-256" in page.casefold() or "sha256" in page.casefold()
    )
    expected = 0
    verified = 0
    for path in source_files:
        if path.suffix.casefold() != ".txt":
            continue
        expected += 1
        digest = hashlib.sha256(path.read_bytes()).hexdigest().casefold()
        verified += int(bool(protocol_text) and digest in protocol_text)
    return {"expected": expected, "verified": verified, "protocol_present": bool(protocol_text)}


def search_corpus(root: Path, query: str, limit: int = 20) -> list[dict]:
    """Search auditable documents without sending the dossier to an external service."""
    terms = [term.casefold() for term in re.findall(r"[\w.-]{2,}", query, re.UNICODE)]
    synonyms = {
        "approval": ("freigabe", "genehmigung"), "approve": ("freigabe", "genehmigt"),
        "payment": ("zahlung", "zahlungslauf"), "vendor": ("kreditor", "lieferant"),
        "supplier": ("lieferant", "kreditor"), "receipt": ("wareneingang", "beleg"),
        "invoice": ("rechnung", "faktura"), "bank": ("bank", "iban", "bankverbindung"),
        "repair": ("reparatur", "instandhaltung"), "asset": ("anlage", "anlagevermögen"),
        "revenue": ("umsatz", "erlös"), "profit": ("gewinn", "jahresüberschuss"),
    }
    terms = list(dict.fromkeys([*terms, *(alias for term in terms for alias in synonyms.get(term, ())) ]))
    if not terms:
        return []
    results: list[dict] = []

    def add(path: Path, locator: dict, text: str) -> None:
        path = source_path(path)
        normalized = text.casefold()
        score = sum(normalized.count(term) for term in terms)
        if not score:
            return
        first = min((normalized.find(term) for term in terms if term in normalized), default=0)
        excerpt = text[max(0, first - 120): first + 360].strip()
        results.append({
            "file": path.relative_to(root).as_posix(),
            "locator": locator,
            "excerpt": excerpt,
            "score": score,
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        })

    for path in root.rglob("*"):
        if not path.is_file() or any(word in path.name.upper() for word in BLOCKED_NAMES):
            continue
        suffix = path.suffix.casefold()
        if suffix == ".pdf":
            for page, text in enumerate(_pdf_pages(path), start=1):
                add(path, {"page": page}, text)
        elif suffix in {".docx", ".pptx"}:
            for text, _, locator in _document_passages(path):
                add(path, locator, text)
        elif suffix == ".xlsx":
            for sheet, rows in _xlsx_rows(path).items():
                for row in rows:
                    add(path, {"sheet": sheet, "row": row["_row"]}, " · ".join(str(value) for value in row["values"] if value not in (None, "")))
        elif suffix in {".csv", ".txt", ".xml", ".json", ".md"} and path.stat().st_size <= 5 * 1024 * 1024:
            try:
                for row, text in enumerate(path.read_text(encoding=_encoding(path)).splitlines(), start=1):
                    add(path, {"row": row}, text)
            except (OSError, UnicodeError):
                pass
        if len(results) > limit * 20:
            results.sort(key=lambda result: result["score"], reverse=True)
            del results[limit * 10:]
    return sorted(results, key=lambda result: (-result["score"], result["file"]))[:limit]


def evidence_context(root: Path, evidence: dict) -> dict:
    """Return neighboring source rows/passages with the cited item marked relevant."""
    root = root.resolve()
    path = (root / evidence["file"]).resolve()
    if root not in path.parents or not path.is_file():
        raise ValueError("Evidence source is unavailable")
    locator = evidence.get("locator") or {}

    def pack(records: list[dict], start: int, end: int) -> dict:
        selected = [record for record in records if start - 1 <= int(record.get("_row", 0)) <= end + 1]
        if len(selected) > 30:
            selected = selected[:15] + selected[-15:]
        columns = list(dict.fromkeys(key for record in selected for key in record if not key.startswith("_")))
        return {
            "columns": columns,
            "rows": [
                {"position": int(record.get("_row", 0)), "relevant": start <= int(record.get("_row", 0)) <= end, "values": {key: record.get(key) for key in columns}}
                for record in selected
            ],
            "source": evidence["file"], "locator": locator,
        }

    start = int(locator.get("row_start") or locator.get("row") or 0)
    end = int(locator.get("row_end") or start)
    suffix = path.suffix.casefold()
    if suffix in {".csv", ".txt"} and start:
        records = []
        if suffix == ".txt":
            records = next((rows for rows in _gdpdu_tables(root).values() if rows and Path(rows[0]["_file"]).resolve() == path), [])
        return pack(records or _csv_rows(path), start, end)

    if suffix == ".xlsx":
        sheet_name = locator.get("sheet")
        sheets = _xlsx_rows(path)
        rows = sheets.get(sheet_name, []) if sheet_name else next(iter(sheets.values()), [])
        if not start and locator.get("range"):
            numbers = [int(value) for value in re.findall(r"\d+", str(locator["range"]))]
            start, end = (min(numbers), max(numbers)) if numbers else (1, 1)
        selected = [row for row in rows if start - 1 <= row["_row"] <= end + 1]
        width = max((len(row["values"]) for row in selected), default=0)
        columns = [f"Column {index + 1}" for index in range(width)]
        return {
            "columns": columns,
            "rows": [{"position": row["_row"], "relevant": start <= row["_row"] <= end, "values": {columns[index]: value for index, value in enumerate(row["values"]) if index < width}} for row in selected],
            "source": evidence["file"], "locator": locator,
        }

    passages = _document_passages(path)
    key = "page" if "page" in locator else "paragraph" if "paragraph" in locator else "slide" if "slide" in locator else None
    current = int(locator.get(key, 1)) if key else 1
    rows = []
    for text, kind, item_locator in passages:
        position = int(item_locator.get(key, item_locator.get(kind, 0)) or 0)
        if current - 1 <= position <= current + 1:
            rows.append({"position": position, "relevant": position == current, "values": {"Passage": text}})
    return {"columns": ["Passage"], "rows": rows or [{"position": current, "relevant": True, "values": {"Passage": evidence.get("excerpt", "")}}], "source": evidence["file"], "locator": locator}


def analyze(
    root: Path,
    supplemental_tables: dict[str, list[dict]] | None = None,
    supplemental_documents: list[dict] | None = None,
) -> dict:
    root = root.resolve()
    tables = _gdpdu_tables(root)
    csvs = _companion_csv(root)
    xlsx = {path.name: _xlsx_rows(path) for path in root.rglob("*.xlsx")}
    structured = {**tables, **csvs, **(supplemental_tables or {})}
    book = EvidenceBook(root)
    findings: list[dict] = []

    ap = _find_rows(structured, {"LIEFERANTENKONTONUMMER", "BUCHUNGSNUMMER", "BUCHUNGSBETRAG", "BUCHUNGSTEXT"})
    vendors = _find_rows(structured, {"LIEFERANTENKONTONUMMER", "LIEFERANTENNAME"})
    gl = _find_rows(structured, {"SACHKONTONUMMER", "BUCHUNGSBETRAG", "DOKUMENT", "BUCHUNGSDATUM", "BENUTZERKENNUNG"})
    assets = _find_rows(structured, {"ANLAGENNUMMER", "ANLAGENBEZEICHNUNG", "ANLAGENGRUPPE"})
    asset_postings = _find_rows(structured, {"ANLAGENNUMMER", "BELEGNUMMER", "BUCHUNGSART"})
    accounts = _find_rows(structured, {"SACHKONTONUMMER", "SACHKONTONAME"})
    receipts = _find_rows(structured, {"KREDITOR", "WARENEINGANG_DATUM", "BETRAG_EUR"})
    changes = _find_rows(structured, {"KONTO", "FELD", "GEAENDERT_VON", "GENEHMIGT_VON"})
    if not gl:
        raise ValueError(
            "No supported general-ledger table found. Expected GDPdU columns "
            "SACHKONTONUMMER, BUCHUNGSBETRAG, DOKUMENT, BUCHUNGSDATUM and BENUTZERKENNUNG."
        )

    input_tax_accounts = {
        str(row.get("SACHKONTONUMMER", ""))
        for row in accounts
        if any(word in str(row.get("SACHKONTONAME", "")).casefold() for word in ("vorsteuer", "input tax"))
    }
    repair_accounts = [
        row for row in accounts
        if any(word in str(row.get("SACHKONTONAME", "")).casefold() for word in ("instandhaltung", "reparatur", "repair", "maintenance"))
    ]
    pnl_accounts = {
        str(row.get("SACHKONTONUMMER", ""))
        for row in accounts
        if str(row.get("KONTENART") or row.get("SACHKONTOTYP") or "").strip().casefold() in {"guv", "p&l", "profit and loss", "income statement"}
    }
    asset_groups = {str(row.get("ANLAGENGRUPPE", "")) for row in assets if row.get("ANLAGENGRUPPE")}

    # Vendor integrity: discover self-approved new vendors, then require incompatible permissions and cash flow.
    permission_path, permission_sheet, permission_rows = _xlsx_table(
        xlsx, {"Benutzer", "Buchen", "Zahlungslauf", "Stammdaten/Kreditor anlegen"}
    )
    if not permission_rows:
        permission_rows = _find_rows(structured, {"Benutzer", "Buchen", "Zahlungslauf", "Stammdaten/Kreditor anlegen"})
        if permission_rows:
            permission_path = Path(permission_rows[0]["_file"])
            permission_sheet = permission_rows[0].get("_sheet")
    permissions = {str(row.get("Benutzer", "")): row for row in permission_rows}
    self_approved = [
        row for row in changes
        if any(term in str(row.get("FELD", "")).casefold() for term in ("neuanlage kreditor", "new vendor", "create vendor"))
        and row.get("GEAENDERT_VON") == row.get("GENEHMIGT_VON")
    ]
    for change in self_approved:
        vendor_id = change.get("KONTO", "")
        user = change.get("GEAENDERT_VON", "")
        rights = permissions.get(user, {})
        incompatible = all(_enabled(rights.get(key)) for key in ("Buchen", "Zahlungslauf", "Stammdaten/Kreditor anlegen"))
        vendor_ap = [row for row in ap if row.get("LIEFERANTENKONTONUMMER") == vendor_id]
        invoices = [row for row in vendor_ap if money(row.get("BUCHUNGSBETRAG")) < 0]
        payments = [row for row in vendor_ap if money(row.get("BUCHUNGSBETRAG")) > 0]
        receipt_matches = [row for row in receipts if row.get("KREDITOR") == vendor_id]
        pairs: list[tuple[dict, dict, int]] = []
        for invoice in invoices:
            invoice_date = _date(invoice.get("BUCHUNGSDATUM"))
            candidates = [
                payment for payment in payments
                if payment.get("BUCHUNGSNUMMER") == invoice.get("BUCHUNGSNUMMER")
                and money(payment.get("BUCHUNGSBETRAG")) == -money(invoice.get("BUCHUNGSBETRAG"))
                and invoice_date
                and _date(payment.get("BUCHUNGSDATUM"))
                and 0 <= (_date(payment.get("BUCHUNGSDATUM")) - invoice_date).days <= 7
            ]
            if candidates:
                payment = min(candidates, key=lambda row: _date(row.get("BUCHUNGSDATUM")))
                pairs.append((invoice, payment, (_date(payment.get("BUCHUNGSDATUM")) - invoice_date).days))
        if not (incompatible and len(pairs) >= 2 and not receipt_matches):
            continue
        master = next((row for row in vendors if row.get("LIEFERANTENKONTONUMMER") == vendor_id), {})
        docs = {invoice.get("BUCHUNGSNUMMER") for invoice, _, _ in pairs}
        gl_lines = [row for row in gl if row.get("DOKUMENT") in docs or row.get("SACHKONTONUMMER") == f"330000-{vendor_id}"]
        booking_users = {str(row.get("BENUTZERKENNUNG", "")) for row in gl_lines}
        if user not in booking_users:
            continue
        net = sum((
            money(row.get("BUCHUNGSBETRAG")) for row in gl_lines
            if (
                str(row.get("BUCHUNGSTYP", "")).casefold() in {"kreditorenrechnung", "vendor invoice", "supplier invoice", "purchase invoice"}
                or str(row.get("SACHKONTONUMMER", "")) in pnl_accounts
            )
            and money(row.get("BUCHUNGSBETRAG")) > 0
            and row.get("SACHKONTONUMMER") not in input_tax_accounts
        ), Decimal("0"))
        tax = sum((money(row.get("BUCHUNGSBETRAG")) for row in gl_lines if row.get("SACHKONTONUMMER") in input_tax_accounts), Decimal("0"))
        gross = sum((money(payment.get("BUCHUNGSBETRAG")) for _, payment, _ in pairs), Decimal("0"))
        lags = [lag for _, _, lag in pairs]
        e_change = book.add(Path(change["_file"]), "row", {"row": change["_row"]}, "Self-approved vendor creation", _record_excerpt(change, ["DATUM", "KONTO", "NAME", "GEAENDERT_VON", "GENEHMIGT_VON"]))
        e_perm = book.add(permission_path, "cell", {"sheet": permission_sheet, "range": f"A{rights['_row']}:J{rights['_row']}"}, "Incompatible permissions", f"{user} can post, run payments, and create vendors.")
        e_ap = book.add(Path(pairs[0][0]["_file"]), "rows", _row_range([row for pair in pairs for row in pair[:2]]), f"{len(pairs)} matched invoice/payment pairs", f"{len(pairs)} exact document-and-amount pairs settled {min(lags)}–{max(lags)} days after invoice.", {"pair_count": len(pairs), "payment_lag_days": lags})
        e_gl = book.add(Path(gl_lines[0]["_file"]), "rows", _row_range(gl_lines), "Net, VAT, cash, and posting-user trail", f"Expense {net}; input VAT {tax}; gross cash {gross}; observed posting users {', '.join(sorted(booking_users))}.")
        e_zero = book.add(Path(receipts[0]["_file"]) if receipts else Path(change["_file"]), "query", {"predicate": f"KREDITOR = {vendor_id}", "population_count": len(receipts), "match_count": 0}, "No matched goods receipt", f"Reproducible filter returned zero matches across {len(receipts)} receipt records.")
        comparator = next((row for row in changes if row.get("KONTO") != vendor_id and any(term in str(row.get("FELD", "")).casefold() for term in ("neuanlage kreditor", "new vendor", "create vendor")) and row.get("GEAENDERT_VON") != row.get("GENEHMIGT_VON")), None)
        counter = []
        if comparator:
            comp_receipts = [row for row in receipts if row.get("KREDITOR") == comparator.get("KONTO")]
            counter.append(book.add(Path(comparator["_file"]), "row", {"row": comparator["_row"]}, "Legitimate new-vendor comparator", f"Independent approval and {len(comp_receipts)} matched receipts for {comparator.get('NAME')}."))
        findings.append({
            "id": "vendor-control-chain" if not any(item["id"].startswith("vendor-control-chain") for item in findings) else f"vendor-control-chain-{vendor_id}",
            "title": f"Probable sham-vendor control chain: {master.get('LIEFERANTENNAME', change.get('NAME'))}",
            "category": "Cash misappropriation",
            "system_status": "investigate",
            "auditor_status": "unreviewed",
            "severity": "critical",
            "confidence": "high",
            "summary": "The same user created and approved the vendor, held posting and payment-run rights, and appears on the linked ledger entries; the complete receipt population contains no match.",
            "amounts": {"net": str(net), "tax": str(tax), "gross": str(gross), "pnl_effect": None},
            "facts": [
                {"label": "Vendor", "value": vendor_id, "evidence_id": e_change},
                {"label": "Creator and approver", "value": user, "evidence_id": e_change},
                {"label": "Gross cash", "value": str(gross), "evidence_id": e_gl, "format": "currency"},
                {"label": "Payment lag", "value": f"{min(lags)}–{max(lags)} days", "evidence_id": e_ap},
                {"label": "Receipt matches", "value": "0", "evidence_id": e_zero},
            ],
            "caveats": ["Absence of a goods receipt alone does not prove consulting services were not delivered.", "Obtain the contract, work product, service acceptance, and beneficiary details."],
            "next_step": "Independently confirm service delivery and bank-account ownership, then inspect the user’s access history.",
            "evidence_ids": [e_change, e_perm, e_ap, e_gl, e_zero],
            "counter_evidence_ids": counter,
            "graph": _graph(
                [("user", user, "user"), ("rights", "Post · create · pay", "permission"), ("vendor", master.get("LIEFERANTENNAME", vendor_id), "vendor"), ("invoices", f"{len(pairs)} invoices", "invoice"), ("payments", f"{len(pairs)} matched payments", "payment"), ("bank", "Cash out", "bank")],
                [("user", "rights", "holds"), ("user", "vendor", "created + approved"), ("vendor", "invoices", "issued"), ("invoices", "payments", "paid rapidly"), ("payments", "bank", "settled")],
            ),
        })

    # Vendor bank-detail changes followed by cash: require failed independent approval and a short, dated link.
    bank_changes = [
        row for row in changes
        if any(term in str(row.get("FELD", "")).casefold() for term in ("iban", "bankverbindung", "bank account", "routing"))
        and (
            row.get("GEAENDERT_VON") == row.get("GENEHMIGT_VON")
            or str(row.get("GENEHMIGT", "")).strip().casefold() in {"no", "nein", "false", "0"}
        )
    ]
    for index, change in enumerate(bank_changes[:5], start=1):
        changed_on = _date(change.get("DATUM"))
        vendor_id = str(change.get("KONTO", ""))
        linked_payments = [
            row for row in ap
            if str(row.get("LIEFERANTENKONTONUMMER", "")) == vendor_id
            and money(row.get("BUCHUNGSBETRAG")) > 0
            and changed_on and _date(row.get("BUCHUNGSDATUM"))
            and 0 <= (_date(row.get("BUCHUNGSDATUM")) - changed_on).days <= 30
            and any(term in str(row.get("BUCHUNGSTEXT", "")).casefold() for term in ("zahlung", "payment", "transfer", "settlement"))
        ]
        if not linked_payments:
            continue
        cash = sum((money(row.get("BUCHUNGSBETRAG")) for row in linked_payments), Decimal("0"))
        e_change = book.add(Path(change["_file"]), "row", {"row": change["_row"]}, "Bank details changed without independent approval", _record_excerpt(change, ["DATUM", "KONTO", "FELD", "WERT_ALT", "WERT_NEU", "GEAENDERT_VON", "GENEHMIGT_VON"]))
        e_cash = book.add(Path(linked_payments[0]["_file"]), "rows", _row_range(linked_payments), "Payments after bank-detail change", f"{len(linked_payments)} payments totaling {cash} EUR were recorded within 30 days of the change.")
        findings.append({
            "id": f"bank-change-before-payment-{index}", "title": f"Unapproved vendor bank change precedes payment: {vendor_id}",
            "category": "Vendor master data / cash diversion", "system_status": "investigate", "auditor_status": "unreviewed", "severity": "critical", "confidence": "high",
            "summary": "A vendor bank-detail change lacked independent approval and was followed by identified supplier payments within a short dated window.",
            "amounts": {"net": None, "tax": None, "gross": str(cash), "pnl_effect": None},
            "facts": [{"label": "Cash after change", "value": str(cash), "evidence_id": e_cash, "format": "currency"}, {"label": "Payments", "value": str(len(linked_payments)), "evidence_id": e_cash}],
            "caveats": ["The timing and control failure do not establish that the replacement bank account is unauthorized."],
            "next_step": "Confirm the change through a known supplier contact, inspect workflow logs, and verify beneficiary ownership with the bank.",
            "evidence_ids": [e_change, e_cash], "counter_evidence_ids": [],
            "graph": _graph([("user", str(change.get("GEAENDERT_VON", "")), "user"), ("vendor", vendor_id, "vendor"), ("bank", "Changed bank details", "account"), ("payment", f"{len(linked_payments)} payments", "payment")], [("user", "vendor", "changed"), ("vendor", "bank", "redirected to"), ("bank", "payment", "received")]),
        })

    # Repairs recorded as new fixed-asset acquisitions.
    repair_ap = [row for row in ap if money(row.get("BUCHUNGSBETRAG")) < 0 and any(word in str(row.get("BUCHUNGSTEXT", "")).lower() for word in REPAIR_WORDS)]
    repair_docs = {row.get("BUCHUNGSNUMMER") for row in repair_ap}
    repair_gl = [row for row in gl if row.get("DOKUMENT") in repair_docs]
    asset_lines = [
        row for row in repair_gl
        if any(
            str(row.get("SACHKONTONUMMER", "")) == group
            or str(row.get("SACHKONTONUMMER", "")).startswith(f"{group}-")
            for group in asset_groups
        )
    ]
    repair_assets = [row for row in assets if any(word in str(row.get("ANLAGENBEZEICHNUNG", "")).lower() for word in REPAIR_WORDS)]
    repair_asset_postings = [row for row in asset_postings if row.get("BELEGNUMMER") in repair_docs]
    if repair_ap and asset_lines and repair_assets and repair_asset_postings:
        net = sum((money(row.get("BUCHUNGSBETRAG")) for row in asset_lines), Decimal("0"))
        tax = sum((money(row.get("BUCHUNGSBETRAG")) for row in repair_gl if row.get("SACHKONTONUMMER") in input_tax_accounts), Decimal("0"))
        gross = abs(sum((money(row.get("BUCHUNGSBETRAG")) for row in repair_ap), Decimal("0")))
        e_ap = book.add(Path(repair_ap[0]["_file"]), "rows", _row_range(repair_ap), "Repair-labelled supplier invoices", "; ".join(row.get("BUCHUNGSTEXT", "") for row in repair_ap))
        e_assets = book.add(Path(repair_assets[0]["_file"]), "rows", _row_range(repair_assets), "Repair descriptions registered as assets", "; ".join(row.get("ANLAGENBEZEICHNUNG", "") for row in repair_assets))
        e_asset_post = book.add(Path(repair_asset_postings[0]["_file"]), "rows", _row_range(repair_asset_postings), "Fixed-asset acquisitions", f"{len(repair_asset_postings)} acquisitions total {net} net.")
        e_gl = book.add(Path(asset_lines[0]["_file"]), "rows", _row_range(repair_gl), "Posted to asset accounts", f"Net {net}; VAT {tax}; gross {gross}.")
        repair_account = repair_accounts[0] if repair_accounts else None
        counter = [book.add(Path(repair_account["_file"]), "row", {"row": repair_account["_row"]}, "Normal repair expense account exists", _record_excerpt(repair_account, ["SACHKONTONUMMER", "SACHKONTONAME"]))] if repair_account else []
        findings.append({
            "id": "capitalized-repairs",
            "title": "Repair-like costs capitalized as fixed assets",
            "category": "Classification / management override",
            "system_status": "investigate",
            "auditor_status": "unreviewed",
            "severity": "high",
            "confidence": "medium",
            "summary": "Repair and overhaul descriptions flow from supplier invoices into newly created fixed assets instead of the repair expense account.",
            "amounts": {"net": str(net), "tax": str(tax), "gross": str(gross), "pnl_effect": str(-net)},
            "facts": [{"label": "Potential adjustment", "value": str(net), "evidence_id": e_gl, "format": "currency"}, {"label": "Items", "value": str(len(repair_ap)), "evidence_id": e_ap}],
            "caveats": ["Major replacements or overhauls can qualify for capitalization depending on scope and policy."],
            "next_step": "Inspect invoice scope, component accounting, capitalization policy, and evidence of increased useful life.",
            "evidence_ids": [e_ap, e_assets, e_asset_post, e_gl],
            "counter_evidence_ids": counter,
            "graph": _graph(
                [("invoice", f"{len(repair_ap)} repair invoices", "invoice"), ("asset", f"{len(asset_lines)} asset debits", "asset"), ("ledger", "Fixed assets", "account"), ("pnl", "Repair expense", "account")],
                [("invoice", "asset", "registered as"), ("asset", "ledger", "capitalized"), ("invoice", "pnl", "not posted")],
            ),
        })

    # Cut-off: January invoice journal with December performance dates and open December receipts.
    january = _find_rows(structured, {"RECHNUNGSNUMMER", "KREDITOR", "FAKTURADATUM", "LEISTUNGSDATUM", "BETRAG_EUR"})
    candidates = [
        row for row in january
        if _date(row.get("LEISTUNGSDATUM"))
        and _date(row.get("FAKTURADATUM"))
        and _date(row.get("FAKTURADATUM")).year == _date(row.get("LEISTUNGSDATUM")).year + 1
    ]
    receipt_pairs: list[tuple[dict, dict]] = []
    for invoice in candidates:
        matches = [
            receipt for receipt in receipts
            if (
                receipt.get("RECHNUNGSNUMMER")
                and receipt.get("RECHNUNGSNUMMER") == invoice.get("RECHNUNGSNUMMER")
            ) or (
                receipt.get("KREDITOR") == invoice.get("KREDITOR")
                and money(receipt.get("BETRAG_EUR")) == money(invoice.get("BETRAG_EUR"))
                and _date(receipt.get("WARENEINGANG_DATUM")) == _date(invoice.get("LEISTUNGSDATUM"))
                and "offen" in str(receipt.get("BEMERKUNG", "")).casefold()
            )
        ]
        if matches:
            receipt_pairs.append((invoice, matches[0]))
    cutoff = [invoice for invoice, _ in receipt_pairs]
    cutoff_docs = {row.get("RECHNUNGSNUMMER") for row in cutoff}
    open_receipts = [receipt for _, receipt in receipt_pairs]
    gl_matches = [row for row in gl if row.get("DOKUMENT") in cutoff_docs or row.get("BELEGNUMMER") in cutoff_docs]
    if cutoff and open_receipts:
        net = sum((money(row.get("BETRAG_EUR")) for row in cutoff), Decimal("0"))
        e_jan = book.add(Path(cutoff[0]["_file"]), "rows", _row_range(cutoff), "January invoices for December performance", f"{len(cutoff)} invoices total {net} with December performance dates.")
        e_receipts = book.add(Path(open_receipts[0]["_file"]), "rows", _row_range(open_receipts), "Open December goods receipts", f"{len(open_receipts)} matching/open receipts were recorded before year-end.")
        e_zero = book.add(Path(gl[0]["_file"]), "query", {"predicate": f"DOKUMENT IN ({', '.join(sorted(cutoff_docs))})", "population_count": len(gl), "match_count": len(gl_matches)}, "No matching current-year posting", f"The complete {len(gl)}-row general ledger contains {len(gl_matches)} matching document IDs.")
        accrual = [row for row in gl if "unfakturierte leistungen" in str(row.get("BUCHUNGSTEXT", "")).lower()]
        counter = []
        if accrual:
            counter.append(book.add(Path(accrual[0]["_file"]), "rows", _row_range(accrual), "Generic accrual exists but is unmapped", "An unbilled-services accrual exists, but carries no supplier or invoice mapping to this population."))
        findings.append({
            "id": "year-end-cutoff",
            "title": "December obligations parked in the next period",
            "category": "Cut-off / unrecorded liabilities",
            "system_status": "exception",
            "auditor_status": "unreviewed",
            "severity": "high",
            "confidence": "high",
            "summary": "Goods were received before year-end, while matching invoices appear only in January and no document-level year-end posting exists.",
            "amounts": {"net": str(net), "tax": None, "gross": None, "pnl_effect": str(-net)},
            "facts": [{"label": "Unrecorded-liability population", "value": str(net), "evidence_id": e_jan, "format": "currency"}, {"label": "Invoices", "value": str(len(cutoff)), "evidence_id": e_jan}],
            "caveats": ["The profit effect depends on whether the goods remained in inventory or were consumed before year-end.", "Do not net the generic accrual without item-level mapping."],
            "next_step": "Trace each receipt into inventory consumption and obtain the year-end accrual detail.",
            "evidence_ids": [e_jan, e_receipts, e_zero],
            "counter_evidence_ids": counter,
            "graph": _graph(
                [("receipt", f"{len(open_receipts)} December receipts", "receipt"), ("close", "Year-end close", "control"), ("invoice", f"{len(cutoff)} January invoices", "invoice"), ("ledger", "No document-level accrual", "journal")],
                [("receipt", "close", "before"), ("close", "invoice", "after"), ("invoice", "ledger", "missing in current year")],
            ),
        })

    # Payment splitting: derive approval threshold from the audit-planning document.
    plan_path = None
    threshold_paragraph = ""
    threshold_kind = "paragraph"
    threshold_locator: dict = {}
    match = None
    passages = (
        (candidate, text, kind, locator)
        for candidate in root.rglob("*")
        if candidate.suffix.casefold() in {".docx", ".pptx", ".pdf", ".txt", ".csv", ".xml", ".md", ".json"}
        for text, kind, locator in _document_passages(candidate)
    )
    vision_passages = (
        (root / item["file"], item["excerpt"], item.get("kind", "vision"), item["locator"])
        for item in supplemental_documents or []
    )
    for candidate, candidate_threshold, candidate_kind, candidate_locator in (*passages, *vision_passages):
        if not re.search(r"Zahlungsfreigaben|payment approval", candidate_threshold, re.I):
            continue
        candidate_match = re.search(r"(?:ab|from|over)\s+([\d.,]+)\s*(?:EUR|€)", candidate_threshold, re.I)
        if candidate_match:
            plan_path, threshold_paragraph, threshold_kind, threshold_locator, match = candidate, candidate_threshold, candidate_kind, candidate_locator, candidate_match
            break
    threshold_text = match.group(1) if match else ""
    threshold = (
        money(re.sub(r"[.,]", "", threshold_text))
        if re.fullmatch(r"\d{1,3}(?:[.,]\d{3})+", threshold_text)
        else money(threshold_text)
    ) if match else None
    groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in ap:
        amount = money(row.get("BUCHUNGSBETRAG"))
        payment_text = str(row.get("BUCHUNGSTEXT", "")).casefold()
        if amount > 0 and any(word in payment_text for word in ("zahlung", "payment", "transfer", "settlement")):
            groups[(row.get("LIEFERANTENKONTONUMMER", ""), row.get("BUCHUNGSDATUM", ""))].append(row)
    split_candidates = [rows for rows in groups.values() if threshold is not None and len(rows) >= 3 and all(threshold * Decimal("0.9") <= money(row.get("BUCHUNGSBETRAG")) < threshold for row in rows)]
    if split_candidates:
        split = max(split_candidates, key=lambda rows: sum(money(row.get("BUCHUNGSBETRAG")) for row in rows))
        total = sum((money(row.get("BUCHUNGSBETRAG")) for row in split), Decimal("0"))
        vendor_id = split[0].get("LIEFERANTENKONTONUMMER", "")
        vendor = next((row for row in vendors if row.get("LIEFERANTENKONTONUMMER") == vendor_id), {})
        e_payments = book.add(Path(split[0]["_file"]), "rows", _row_range(split), "Same-day payments below approval threshold", "; ".join(f"{row.get('BUCHUNGSDATUM')}: {row.get('BUCHUNGSBETRAG')} EUR" for row in split))
        e_policy = book.add(plan_path, threshold_kind, threshold_locator, "Payment approval policy", threshold_paragraph)
        findings.append({
            "id": "threshold-splitting",
            "title": f"Payments structured below approval threshold: {vendor.get('LIEFERANTENNAME', vendor_id)}",
            "category": "Control circumvention",
            "system_status": "investigate",
            "auditor_status": "unreviewed",
            "severity": "medium",
            "confidence": "high",
            "summary": "Several same-day payments to one supplier fall immediately below the sourced two-approval threshold.",
            "amounts": {"net": None, "tax": None, "gross": str(total), "pnl_effect": None},
            "facts": [{"label": "Combined payments", "value": str(total), "evidence_id": e_payments, "format": "currency"}, {"label": "Approval threshold", "value": str(threshold), "evidence_id": e_policy, "format": "currency"}],
            "caveats": ["The pattern proves a control indicator, not intent or financial misstatement."],
            "next_step": "Inspect payment proposals, approver logs, invoice allocation, and correspondence for the payment date.",
            "evidence_ids": [e_payments, e_policy],
            "counter_evidence_ids": [],
            "graph": _graph(
                [("vendor", vendor.get("LIEFERANTENNAME", vendor_id), "vendor"), ("payments", f"{len(split)} same-day payments", "payment"), ("threshold", "Two-approval threshold", "control"), ("bank", "Combined cash out", "bank")],
                [("vendor", "payments", "received"), ("payments", "threshold", "each below"), ("payments", "bank", "combined")],
            ),
        })

    # Generic duplicate and overpayment tests: exact identifiers only, to protect precision.
    invoice_groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    document_groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in ap:
        vendor_id = str(row.get("LIEFERANTENKONTONUMMER", ""))
        external_id = str(row.get("BELEGNUMMER") or row.get("RECHNUNGSNUMMER") or "").strip()
        document_id = str(row.get("BUCHUNGSNUMMER", "")).strip()
        if external_id and money(row.get("BUCHUNGSBETRAG")) < 0:
            invoice_groups[(vendor_id, external_id)].append(row)
        if document_id:
            document_groups[(vendor_id, document_id)].append(row)
    duplicates = [
        rows for rows in invoice_groups.values()
        if len(rows) > 1
        and len({row.get("BUCHUNGSNUMMER") for row in rows}) > 1
        and len({abs(money(row.get("BUCHUNGSBETRAG"))) for row in rows}) == 1
    ]
    for index, rows in enumerate(duplicates[:5], start=1):
        duplicated_amount = abs(money(rows[0].get("BUCHUNGSBETRAG")))
        vendor_id = str(rows[0].get("LIEFERANTENKONTONUMMER", ""))
        external_id = str(rows[0].get("BELEGNUMMER", ""))
        e_duplicate = book.add(
            Path(rows[0]["_file"]), "rows", _row_range(rows), "Duplicate supplier invoice identifier",
            f"External invoice {external_id} appears {len(rows)} times for vendor {vendor_id} at {duplicated_amount} EUR.",
        )
        findings.append({
            "id": f"duplicate-supplier-invoice-{index}",
            "title": f"Supplier invoice posted more than once: {external_id}",
            "category": "Duplicate disbursement / expense overstatement",
            "system_status": "exception", "auditor_status": "unreviewed", "severity": "high", "confidence": "high",
            "summary": "The same non-empty external invoice identifier and amount were posted under different accounting documents for one supplier.",
            "amounts": {"net": str(duplicated_amount), "tax": None, "gross": None, "pnl_effect": str(duplicated_amount)},
            "facts": [{"label": "Duplicate amount", "value": str(duplicated_amount), "evidence_id": e_duplicate, "format": "currency"}, {"label": "Occurrences", "value": str(len(rows)), "evidence_id": e_duplicate}],
            "caveats": ["A deliberate split posting can reuse a reference, but exact equal amounts under different documents require explanation."],
            "next_step": "Inspect both invoice images, reversal history, payment settlement, and supplier confirmation.",
            "evidence_ids": [e_duplicate], "counter_evidence_ids": [],
            "graph": _graph([("invoice", external_id, "invoice"), ("postings", f"{len(rows)} postings", "journal"), ("vendor", vendor_id, "vendor")], [("invoice", "postings", "reused as"), ("postings", "vendor", "charged to")]),
        })
    overpayments = []
    for rows in document_groups.values():
        invoiced = abs(sum((money(row.get("BUCHUNGSBETRAG")) for row in rows if money(row.get("BUCHUNGSBETRAG")) < 0), Decimal("0")))
        paid = sum((money(row.get("BUCHUNGSBETRAG")) for row in rows if money(row.get("BUCHUNGSBETRAG")) > 0), Decimal("0"))
        if invoiced > 0 and paid - invoiced > Decimal("0.01"):
            overpayments.append((rows, paid - invoiced, invoiced, paid))
    for index, (rows, excess, invoiced, paid) in enumerate(sorted(overpayments, key=lambda item: item[1], reverse=True)[:5], start=1):
        document_id = str(rows[0].get("BUCHUNGSNUMMER", ""))
        vendor_id = str(rows[0].get("LIEFERANTENKONTONUMMER", ""))
        e_overpay = book.add(Path(rows[0]["_file"]), "rows", _row_range(rows), "Payment exceeds matched invoice", f"Document {document_id}: invoiced {invoiced}; paid {paid}; excess {excess} EUR.")
        findings.append({
            "id": f"supplier-overpayment-{index}", "title": f"Supplier payment exceeds matched invoice: {document_id}",
            "category": "Cash disbursement", "system_status": "exception", "auditor_status": "unreviewed", "severity": "high", "confidence": "high",
            "summary": "Positive supplier-ledger settlements exceed the negative invoice amount under the same accounting document.",
            "amounts": {"net": None, "tax": None, "gross": str(excess), "pnl_effect": None},
            "facts": [{"label": "Excess cash", "value": str(excess), "evidence_id": e_overpay, "format": "currency"}],
            "caveats": ["The excess may be a deposit or on-account payment if supported by separate documentation."],
            "next_step": "Trace the excess to bank settlement, subsequent allocation, refund, and supplier statement.",
            "evidence_ids": [e_overpay], "counter_evidence_ids": [],
            "graph": _graph([("invoice", f"Invoice {invoiced}", "invoice"), ("payment", f"Paid {paid}", "payment"), ("excess", f"Excess {excess}", "bank")], [("invoice", "payment", "matched to"), ("payment", "excess", "exceeds by")]),
        })

    # Premature revenue: require an invoice date and explicit performance date in different fiscal years.
    sales_journal = _find_rows(structured, {"RECHNUNGSNUMMER", "DEBITOR", "FAKTURADATUM", "LEISTUNGSDATUM", "BETRAG_EUR"})
    premature = [
        row for row in sales_journal
        if _date(row.get("FAKTURADATUM")) and _date(row.get("LEISTUNGSDATUM"))
        and _date(row.get("LEISTUNGSDATUM")).year > _date(row.get("FAKTURADATUM")).year
    ]
    if premature:
        amount = sum((money(row.get("BETRAG_EUR")) for row in premature), Decimal("0"))
        e_revenue = book.add(Path(premature[0]["_file"]), "rows", _row_range(premature), "Invoices precede next-period performance", f"{len(premature)} invoices total {amount} EUR; each explicit performance date falls in a later fiscal year.")
        findings.append({
            "id": "premature-revenue", "title": "Revenue invoiced before next-period performance", "category": "Revenue recognition / cut-off",
            "system_status": "exception", "auditor_status": "unreviewed", "severity": "high", "confidence": "high",
            "summary": "Invoice dates fall in the reporting period while the source journal records performance dates in the following fiscal year.",
            "amounts": {"net": str(amount), "tax": None, "gross": None, "pnl_effect": str(-amount)},
            "facts": [{"label": "Revenue population", "value": str(amount), "evidence_id": e_revenue, "format": "currency"}, {"label": "Invoices", "value": str(len(premature)), "evidence_id": e_revenue}],
            "caveats": ["Contract terms may support over-time recognition; the journal dates alone do not resolve that accounting judgment."],
            "next_step": "Inspect contracts, dispatch evidence, acceptance milestones, credit notes, and the subsequent-period ledger.",
            "evidence_ids": [e_revenue], "counter_evidence_ids": [],
            "graph": _graph([("invoice", f"{len(premature)} current-year invoices", "invoice"), ("close", "Fiscal close", "control"), ("service", "Next-year performance", "receipt")], [("invoice", "close", "before"), ("close", "service", "before")]),
        })

    # Keep contradictory master data visible without promoting it as fraud.
    signals: list[dict] = []
    shareholder_rows = _find_rows(structured, {"NAME", "VERHAELTNIS", "BEMERKUNG"})
    for row in shareholder_rows:
        account_match = re.search(r"Kreditor\s+(\d+)", str(row.get("BEMERKUNG", "")), re.I)
        if not account_match:
            continue
        vendor = next((item for item in vendors if item.get("LIEFERANTENKONTONUMMER") == account_match.group(1)), None)
        if vendor and str(row.get("NAME", "")).casefold() not in str(vendor.get("LIEFERANTENNAME", "")).casefold():
            e_related = book.add(Path(row["_file"]), "row", {"row": row["_row"]}, "Related-party account mapping", _record_excerpt(row, ["NAME", "VERHAELTNIS", "BEMERKUNG"]))
            e_vendor = book.add(Path(vendor["_file"]), "row", {"row": vendor["_row"]}, "Vendor master identity", _record_excerpt(vendor, ["LIEFERANTENKONTONUMMER", "LIEFERANTENNAME", "LIEFERANTENUSTIDNR"]))
            signals.append({"id": "related-party-mapping", "status": "data_quality", "title": "Related-party account maps to different legal names", "evidence_ids": [e_related, e_vendor], "disposition": "Do not allege concealment; reconcile the legal-entity mapping."})

    files = [path for path in root.rglob("*") if path.is_file() and ".proofline." not in path.name and not any(word in path.name.upper() for word in BLOCKED_NAMES)]
    row_counts = {name: len(rows) for name, rows in tables.items()} | {name: len(rows) for name, rows in csvs.items()}
    manifest = []
    for path in sorted(files):
        records = row_counts.get(path.name)
        if path.suffix.lower() == ".pdf":
            records = len(_pdf_pages(path))
        elif path.suffix.lower() in {".docx", ".pptx"}:
            records = len(_document_passages(path))
        elif path.suffix.lower() == ".xlsx":
            records = sum(len(rows) for rows in xlsx.get(path.name, {}).values())
        manifest.append({
            "path": path.relative_to(root).as_posix(),
            "bytes": path.stat().st_size,
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            "kind": path.suffix.lower().lstrip(".") or "file",
            "records": records,
            "status": "parsed" if path.suffix.lower() in {".txt", ".csv", ".xml", ".xlsx", ".pdf", ".docx", ".pptx", ".md", ".json", ".png", ".jpg", ".jpeg", ".webp"} or any(path.parent.glob(f"{path.name}.proofline.*")) else "preserved",
        })
    trial_path, _, _ = _xlsx_table(xlsx, {"Konto", "Bezeichnung", "Kontenart"})
    trial = _trial_balance(trial_path)
    gl_total = sum((money(row.get("BUCHUNGSBETRAG")) for row in gl), Decimal("0"))
    source_files = sorted({Path(rows[0]["_file"]) for rows in tables.values() if rows})
    hashes = _hash_verification(root, source_files)
    if gl_total != 0:
        integrity = "attention"
    elif hashes["protocol_present"] and hashes["expected"] and hashes["verified"] == hashes["expected"]:
        integrity = "verified"
    else:
        integrity = "balanced"
    reported = trial["reported_profit"]
    proposed_adjustment = sum((money(finding.get("amounts", {}).get("pnl_effect")) for finding in findings), Decimal("0"))
    reported_evidence = book.add(
        trial_path,
        "cell" if trial.get("reported_cell") else "calculation",
        {"sheet": trial.get("reported_sheet"), "range": trial.get("reported_cell")} if trial.get("reported_cell") else {"sheet": trial.get("sheet"), "procedure": "sum P&L balances"},
        "Reported draft profit",
        f"Jahresüberschuss laut Entwurf: {reported} EUR",
    ) if trial_path and reported is not None else None
    findings.sort(key=lambda finding: {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(finding["severity"], 4))
    claim_count = sum(len(finding.get("facts", [])) + 1 for finding in findings)
    cited_claims = sum(
        int(bool(finding.get("evidence_ids")) and all(eid in book.items for eid in finding["evidence_ids"]))
        + sum(int(fact.get("evidence_id") in book.items) for fact in finding.get("facts", []))
        for finding in findings
    )
    unsupported_claims = max(0, claim_count - cited_claims)
    years = [parsed.year for row in gl if (parsed := _date(row.get("BUCHUNGSDATUM")))]
    fiscal_year = max(set(years), key=years.count) if years else None
    return {
        "run": {
            "id": "sample",
            "status": "complete",
            "engagement": root.name,
            "fiscal_year": fiscal_year,
            "files": len(files),
            "row_count": sum(row_counts.values()),
            "integrity": integrity,
            "reported_profit": str(reported) if reported is not None else None,
            "proposed_adjusted_profit": str(reported + proposed_adjustment) if reported is not None else None,
            "reported_profit_evidence_id": reported_evidence,
        },
        "manifest": manifest,
        "findings": findings,
        "signals": signals,
        "evidence": book.items,
        "calculations": ([{
            "id": "profit-bridge",
            "expression": "reported_profit + capitalized_repairs_adjustment + cutoff_adjustment",
            "inputs": [reported_evidence, *[eid for finding in findings if finding["id"] in {"capitalized-repairs", "year-end-cutoff"} for eid in finding["evidence_ids"]]],
            "output": str(reported + proposed_adjustment),
        }] if reported is not None else []),
        "metrics": {
            "general_ledger_rows": len(gl),
            "general_ledger_balance": str(gl_total),
            "formula_count": trial["formula_count"],
            "formula_resolved": trial["formula_resolved"],
            "citation_coverage": round(100 * cited_claims / claim_count) if claim_count else 0,
            "promoted_findings": len(findings),
            "unsupported_claims": unsupported_claims,
            "cleared_signals": len(signals),
            "hashes_expected": hashes["expected"],
            "hashes_verified": hashes["verified"],
        },
    }
